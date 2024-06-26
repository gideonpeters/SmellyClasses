import openpyxl
import os

class ExcelProcessor:
    def read_excel(self, file_name):
        if not file_name:
            return None
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(tuple(row))
        workbook.close()
        return data

    def write_excel(self, data, save_file_name):
        if not save_file_name:
            return 0
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in data:
            sheet.append(row)
        workbook.save(save_file_name)
        workbook.close()
        return 1

    def process_excel_data(self, N, file_name):
        if not file_name:
            return 0, None
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        processed_data = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                row += (row[N],)
            else:
                row += (row[N],)
            processed_data.append(row)
        output_file = 'processed_' + file_name
        workbook.save(output_file)
        workbook.close()
        return 1, output_file
